#!/usr/bin/env python3
"""Build an Excel performance-curve ("knee") chart comparing the auth service
running as 1 instance (one.txt) vs 3 instances (three.txt).

Data was extracted from the k6 summaries produced by run_login_test_remote.sh
against the AWS NLB. Each VU level was a separate 30s k6 run.

Run:  python3 tests/perf/build_knee_chart.py
Out:  tests/perf/auth-scaling-knee.xlsx
"""
from openpyxl import Workbook
from openpyxl.chart import LineChart, ScatterChart, Reference, Series
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# VUs, avg http_req_duration (ms), http_req_failed (%)
ONE = [
    (1,   488.71,  4.76),
    (50,  9140.0,  0.58),
    (100, 16550.0, 0.45),
    (200, 23460.0, 23.52),
    (300, 27850.0, 54.18),
]
THREE = [
    (1,   564.78, 5.00),
    (50,  2650.0, 0.23),
    (100, 5760.0, 0.21),
    (200, 11370.0, 0.18),
    (300, 15270.0, 0.15),
]

wb = Workbook()
ws = wb.active
ws.title = "Data"

bold = Font(bold=True)
white_bold = Font(bold=True, color="FFFFFF")
center = Alignment(horizontal="center")
hdr_fill = PatternFill("solid", fgColor="1F4E78")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ---- Title ----
ws["A1"] = "Auth Service Login Performance — 1 vs 3 instances (k6 via AWS NLB)"
ws["A1"].font = Font(bold=True, size=13)
ws.merge_cells("A1:E1")

# ---- Header row ----
headers = ["VUs (concurrent users)",
           "1 inst — avg (ms)", "1 inst — failed %",
           "3 inst — avg (ms)", "3 inst — failed %"]
row0 = 3
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=row0, column=c, value=h)
    cell.font = white_bold
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    cell.border = border

# ---- Data rows ----
for i, ((vu, avg1, fail1), (_, avg3, fail3)) in enumerate(zip(ONE, THREE)):
    r = row0 + 1 + i
    vals = [vu, round(avg1, 1), fail1, round(avg3, 1), fail3]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.alignment = center
        cell.border = border

last = row0 + len(ONE)  # last data row

# Highlight the knee region for the 1-instance run (failures explode at 200 VUs)
knee_fill = PatternFill("solid", fgColor="FCE4D6")
for c in range(1, 6):
    ws.cell(row=row0 + 4, column=c).fill = knee_fill  # VUs=200 row

# Column widths
for col, w in zip("ABCDE", [22, 18, 16, 18, 16]):
    ws.column_dimensions[col].width = w

# Notes
note_r = last + 2
ws.cell(row=note_r, column=1, value="Notes:").font = bold
notes = [
    "Each row is a separate 30s k6 run at a fixed VU level (no ramp).",
    "avg = http_req_duration (avg);  failed % = http_req_failed.",
    "1-instance KNEE ~200 VUs: failures jump 0.45% -> 23.5% -> 54.2% and latency keeps climbing.",
    "3-instance run stays < 1% failures through 300 VUs and ~2-3x lower latency: no knee reached.",
    "The VU=1 ~5% 'failed' is a single cold-start request (1/20-21); ignore as warm-up noise.",
]
for i, n in enumerate(notes, start=1):
    ws.cell(row=note_r + i, column=1, value="• " + n)
    ws.merge_cells(start_row=note_r + i, start_column=1, end_row=note_r + i, end_column=5)

# ========== Chart 1: Response-time performance curve (the knee) ==========
chart = ScatterChart()
chart.title = "Performance Curve (knee): avg response time vs concurrent users"
chart.style = 2
chart.x_axis.title = "Concurrent users (VUs)"
chart.y_axis.title = "Avg response time (ms)"
chart.height = 11
chart.width = 22
chart.x_axis.delete = False
chart.y_axis.delete = False

xref = Reference(ws, min_col=1, min_row=row0 + 1, max_row=last)

y1 = Reference(ws, min_col=2, min_row=row0, max_row=last)
s1 = Series(y1, xref, title_from_data=True)
s1.marker.symbol = "circle"
s1.marker.size = 7
s1.graphicalProperties.line.solidFill = "C00000"
s1.graphicalProperties.line.width = 28000
chart.series.append(s1)

y3 = Reference(ws, min_col=4, min_row=row0, max_row=last)
s3 = Series(y3, xref, title_from_data=True)
s3.marker.symbol = "diamond"
s3.marker.size = 7
s3.graphicalProperties.line.solidFill = "2E7D32"
s3.graphicalProperties.line.width = 28000
chart.series.append(s3)

ws.add_chart(chart, "G3")

# ========== Chart 2: Error rate vs load ==========
chart2 = ScatterChart()
chart2.title = "Error rate (http_req_failed %) vs concurrent users"
chart2.style = 2
chart2.x_axis.title = "Concurrent users (VUs)"
chart2.y_axis.title = "Failed requests (%)"
chart2.height = 9
chart2.width = 22
chart2.x_axis.delete = False
chart2.y_axis.delete = False

f1 = Reference(ws, min_col=3, min_row=row0, max_row=last)
sf1 = Series(f1, xref, title_from_data=True)
sf1.marker.symbol = "circle"; sf1.marker.size = 7
sf1.graphicalProperties.line.solidFill = "C00000"
sf1.graphicalProperties.line.width = 28000
chart2.series.append(sf1)

f3 = Reference(ws, min_col=5, min_row=row0, max_row=last)
sf3 = Series(f3, xref, title_from_data=True)
sf3.marker.symbol = "diamond"; sf3.marker.size = 7
sf3.graphicalProperties.line.solidFill = "2E7D32"
sf3.graphicalProperties.line.width = 28000
chart2.series.append(sf3)

ws.add_chart(chart2, "G26")

out = "tests/perf/auth-scaling-knee.xlsx"
wb.save(out)
print("wrote", out)
