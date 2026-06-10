#!/usr/bin/env python3
"""Genera docs/prototype4_cache_resultados.xlsx con la comparación de desempeño
del patrón Cache-Aside (Chat service), cache OFF vs ON, medido in-cluster
(k6 -> gateway:8080). Barrido de concurrencia hasta 1000 VUs."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

# Barrido de saturación (DURATION=30s/nivel).
# (VUs, avg, med, p90, p95, max_s, reqps, http_reqs, hit_rate%)
OFF = [
    (100,  17.12, 13.38,  31.50,  38.96, 0.427,  813.12, 25529, 0.0),
    (200, 120.74, 78.12, 265.15, 391.58, 1.030,  860.45, 27084, 0.0),
    (500, 278.63, 231.44, 465.91, 555.56, 2.610, 1250.07, 39376, 0.0),
    (1000, 833.53, 609.18, 1690.0, 2220.0, 4.040, 1013.35, 32142, 0.0),
]
ON = [
    (100,   8.99,  5.42,  13.68,  19.47, 0.509,  873.43, 27444, 99.80),
    (200,  83.12, 32.87, 235.68, 346.66, 0.808, 1023.49, 32551, 99.77),
    (500, 173.76, 119.74, 369.65, 502.14, 1.450, 1708.54, 53892, 99.55),
    (1000, 630.61, 358.16, 1510.0, 1920.0, 5.410, 1262.07, 41100, 98.52),
]

NAVY = "1F3864"; BLUE = "2E5496"; LGREY = "D9E1F2"
GREEN = "C6EFCE"; GREEN_FT = "006100"; WHITE = "FFFFFF"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()

# ----------------------------------------------------------------- Sheet 1
ws = wb.active
ws.title = "Comparación"
ws["A1"] = "Prototipo 4 — Escenario 2 (Desempeño): Patrón Cache-Aside (TTL) en el Chat Service"
ws["A1"].font = Font(bold=True, size=14, color=NAVY); ws.merge_cells("A1:I1")
ws["A2"] = ("GET /channels/{id}/messages · medición in-cluster (k6 → gateway:8080, 30 s/nivel) · "
            "barrido de concurrencia 100→1000 VUs · Redis dedicado, TTL 5 s, clave por canal, tras CheckPerm")
ws["A2"].font = Font(italic=True, size=10, color="595959"); ws.merge_cells("A2:I2")

hdr = ["VUs", "p95 SIN caché (ms)", "p95 CON caché (ms)", "Mejora p95",
       "req/s SIN caché", "req/s CON caché", "Mejora req/s", "Hit-rate CON caché", "Errores"]
r0 = 4
for c, h in enumerate(hdr, start=1):
    cell = ws.cell(row=r0, column=c, value=h)
    cell.font = Font(bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

for i, (off, on) in enumerate(zip(OFF, ON)):
    r = r0 + 1 + i
    p95_off, p95_on = off[4], on[4]
    rps_off, rps_on = off[6], on[6]
    vals = [off[0], p95_off, p95_on, 1 - p95_on / p95_off,
            rps_off, rps_on, rps_on / rps_off - 1, on[8] / 100.0, 0.0]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = border; cell.alignment = center
        if c in (4, 7, 8, 9):
            cell.number_format = "0.0%"
        elif c in (5, 6):
            cell.number_format = "0"
        elif c != 1:
            cell.number_format = "0.00"
        if i % 2 == 0:
            cell.fill = PatternFill("solid", fgColor=LGREY)
    for c in (4, 7):
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=GREEN)
        ws.cell(row=r, column=c).font = Font(bold=True, color=GREEN_FT)

for c, w in enumerate([8, 17, 17, 11, 15, 15, 11, 16, 9], start=1):
    ws.column_dimensions[get_column_letter(c)].width = w

rc = r0 + len(OFF) + 2
ws.cell(row=rc, column=1, value="Hallazgo:").font = Font(bold=True, color=NAVY)
concl = ("Con la caché habilitada, ~99–99.8% de las lecturas se sirven desde Redis (Cassandra recibe "
         "< 1%). Bajo baja concurrencia el beneficio es la LATENCIA (p95 −40–50%). Al subir la carga "
         "el servidor sin caché se satura por Cassandra: el throughput tope sube de ~1250 req/s "
         "(OFF) a ~1708 req/s (ON, +37% @ 500 VUs) y la latencia se mantiene menor con caché en todo "
         "el rango. A 1000 VUs ambos saturan (1 pod de chat, Cassandra RF=1), pero la caché sostiene "
         "más req/s (1262 vs 1013) y menor p95 (1.92 s vs 2.22 s). 0% de errores en todo el barrido.")
ws.cell(row=rc, column=2, value=concl).alignment = left
ws.merge_cells(start_row=rc, start_column=2, end_row=rc + 3, end_column=9)

# Charts
chart1 = BarChart(); chart1.type = "col"
chart1.title = "Latencia p95: SIN vs CON caché (ms)"
chart1.y_axis.title = "p95 (ms)"; chart1.x_axis.title = "VUs"
d1 = Reference(ws, min_col=2, max_col=3, min_row=r0, max_row=r0 + len(OFF))
cats = Reference(ws, min_col=1, min_row=r0 + 1, max_row=r0 + len(OFF))
chart1.add_data(d1, titles_from_data=True); chart1.set_categories(cats)
chart1.height = 8; chart1.width = 14
ws.add_chart(chart1, f"A{rc + 6}")

chart2 = BarChart(); chart2.type = "col"
chart2.title = "Throughput: SIN vs CON caché (req/s)"
chart2.y_axis.title = "req/s"; chart2.x_axis.title = "VUs"
d2 = Reference(ws, min_col=5, max_col=6, min_row=r0, max_row=r0 + len(OFF))
chart2.add_data(d2, titles_from_data=True); chart2.set_categories(cats)
chart2.height = 8; chart2.width = 14
ws.add_chart(chart2, f"F{rc + 6}")

# ----------------------------------------------------------------- Sheet 2
ws2 = wb.create_sheet("Métricas crudas k6")
cols = ["Config", "VUs", "avg (ms)", "med (ms)", "p90 (ms)", "p95 (ms)", "max (s)",
        "req/s", "http_reqs", "cache_hit_rate", "http_req_failed"]
for c, h in enumerate(cols, start=1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = Font(bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor=NAVY); cell.alignment = center; cell.border = border
r = 2
for label, rows in (("SIN caché", OFF), ("CON caché", ON)):
    for row in rows:
        vals = [label, row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8] / 100.0, 0.0]
        for c, v in enumerate(vals, start=1):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.border = border
            cell.alignment = center if c > 1 else Alignment(horizontal="left", vertical="center")
            if c in (10, 11):
                cell.number_format = "0.00%"
            elif 3 <= c <= 6:
                cell.number_format = "0.00"
            elif c == 7:
                cell.number_format = "0.000"
            elif c == 8:
                cell.number_format = "0.00"
        fill = LGREY if label == "SIN caché" else GREEN
        ws2.cell(row=r, column=1).fill = PatternFill("solid", fgColor=fill)
        r += 1
for c, w in enumerate([11, 6, 9, 9, 9, 9, 8, 9, 11, 14, 14], start=1):
    ws2.column_dimensions[get_column_letter(c)].width = w
ws2.cell(row=r + 1, column=1,
         value=("Entorno: EKS us-east-1 · k6 dentro del clúster (Job, 500m CPU) → Service gateway:8080 · "
                "RATE_LIMIT_ENABLED=false solo durante la medición · DURATION=30s por nivel · "
                "chat: 1 réplica · Cassandra RF=1 · ~30 mensajes sembrados por canal · "
                "Nota baja concurrencia (20s): p95 10VUs 10.6→6.2ms, 50VUs 16.7→8.5ms."))
ws2.cell(row=r + 1, column=1).font = Font(italic=True, size=9, color="595959")
ws2.merge_cells(start_row=r + 1, start_column=1, end_row=r + 2, end_column=11)

out = "docs/prototype4_cache_resultados.xlsx"
wb.save(out)
print("escrito:", out)
